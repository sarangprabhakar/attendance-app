import os
import sqlite3
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session
from functools import wraps
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = 'attendance_secret_key_2024'

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
DB_PATH       = os.path.join(BASE_DIR, 'attendance.db')
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# ── DB HELPERS ────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_db():
    conn = get_db()
    cur  = conn.cursor()
    cur.executescript("""
        CREATE TABLE IF NOT EXISTS students (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            regno      TEXT UNIQUE NOT NULL,
            name       TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS subjects (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        );
        CREATE TABLE IF NOT EXISTS attendance_records (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            subject_id INTEGER NOT NULL,
            date       TEXT NOT NULL,
            status     TEXT NOT NULL,
            uploaded_at TEXT DEFAULT (datetime('now')),
            UNIQUE (student_id, subject_id, date),
            FOREIGN KEY (student_id) REFERENCES students(id),
            FOREIGN KEY (subject_id) REFERENCES subjects(id)
        );
        CREATE TABLE IF NOT EXISTS users (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            username      TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role          TEXT DEFAULT 'viewer',
            created_at    TEXT DEFAULT (datetime('now'))
        );
    """)
    cur.execute("SELECT COUNT(*) FROM users")
    if cur.fetchone()[0] == 0:
        cur.execute(
            "INSERT INTO users (username, password_hash, role) VALUES (?,?,?)",
            ('admin', generate_password_hash('admin123'), 'admin')
        )
        print("✅ Default admin — username: admin / password: admin123")
    conn.commit()
    conn.close()

# ── AUTH DECORATORS ───────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to continue.', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to continue.', 'danger')
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('Admin access required.', 'danger')
            return redirect(url_for('report'))
        return f(*args, **kwargs)
    return decorated

# ── DB OPERATIONS ─────────────────────────────────────────────────────────────

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_or_create_subject(cur, name):
    cur.execute("SELECT id FROM subjects WHERE name=?", (name,))
    row = cur.fetchone()
    if row:
        return row['id']
    cur.execute("INSERT INTO subjects (name) VALUES (?)", (name,))
    return cur.lastrowid

def upsert_student(cur, regno, name):
    cur.execute("SELECT id FROM students WHERE regno=?", (regno,))
    row = cur.fetchone()
    if row:
        cur.execute("UPDATE students SET name=? WHERE regno=?", (name, regno))
        return row['id']
    cur.execute("INSERT INTO students (regno, name) VALUES (?,?)", (regno, name))
    return cur.lastrowid

def parse_and_store(filepath, subject_name):
    wb      = load_workbook(filepath, data_only=True)
    results = {'inserted': 0, 'updated': 0, 'errors': []}
    conn    = get_db()
    cur     = conn.cursor()
    subject_id = get_or_create_subject(cur, subject_name)

    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers      = [str(h).strip() if h is not None else '' for h in rows[0]]
        regno_idx    = next((i for i, h in enumerate(headers) if 'reg' in h.lower() or 'roll' in h.lower()), 0)
        name_idx     = next((i for i, h in enumerate(headers) if 'name' in h.lower()), 1)
        date_indices = [i for i in range(len(headers)) if i not in (regno_idx, name_idx)]

        date_map = {}
        for i in date_indices:
            raw = rows[0][i]
            if isinstance(raw, (datetime, date)):
                date_map[i] = raw.date().isoformat() if isinstance(raw, datetime) else raw.isoformat()
            else:
                h = headers[i]
                if not h or h.lower() == 'none':
                    continue
                parsed = None
                for fmt in ('%Y-%m-%d','%d-%m-%Y','%d/%m/%Y','%m/%d/%Y','%d-%b-%Y','%d %b %Y'):
                    try:
                        parsed = datetime.strptime(str(h), fmt).date().isoformat()
                        break
                    except ValueError:
                        continue
                if parsed:
                    date_map[i] = parsed
                else:
                    results['errors'].append(f"Can't parse date: {h}")

        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            regno = str(row[regno_idx]).strip() if row[regno_idx] is not None else ''
            name  = str(row[name_idx]).strip()  if row[name_idx]  is not None else ''
            if not regno or regno.lower() == 'none':
                continue
            try:
                student_id = upsert_student(cur, regno, name)
            except Exception as e:
                results['errors'].append(f"Student {regno}: {e}")
                continue

            for i, date_val in date_map.items():
                if i >= len(row):
                    continue
                status = str(row[i]).strip().upper() if row[i] is not None else ''
                if status not in ('P', 'A'):
                    continue
                try:
                    cur.execute("""
                        SELECT id FROM attendance_records
                        WHERE student_id=? AND subject_id=? AND date=?
                    """, (student_id, subject_id, date_val))
                    existing = cur.fetchone()
                    cur.execute("""
                        INSERT INTO attendance_records (student_id, subject_id, date, status)
                        VALUES (?,?,?,?)
                        ON CONFLICT(student_id, subject_id, date)
                        DO UPDATE SET status=excluded.status, uploaded_at=datetime('now')
                    """, (student_id, subject_id, date_val, status))
                    if existing:
                        results['updated'] += 1
                    else:
                        results['inserted'] += 1
                except Exception as e:
                    results['errors'].append(str(e))

    conn.commit()
    conn.close()
    return results

# ── AUTH ROUTES ───────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET','POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','')
        conn = get_db()
        cur  = conn.cursor()
        cur.execute("SELECT * FROM users WHERE username=?", (username,))
        user = cur.fetchone()
        conn.close()
        if user and check_password_hash(user['password_hash'], password):
            session['user_id']  = user['id']
            session['username'] = user['username']
            session['role']     = user['role']
            flash(f"Welcome, {user['username']}! 👋", 'success')
            return redirect(url_for('index'))
        flash('Invalid username or password.', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully.', 'success')
    return redirect(url_for('login'))

# ── USER MANAGEMENT ───────────────────────────────────────────────────────────

@app.route('/admin/users')
@admin_required
def manage_users():
    conn  = get_db()
    cur   = conn.cursor()
    cur.execute("SELECT id, username, role, created_at FROM users ORDER BY created_at")
    users = cur.fetchall()
    conn.close()
    return render_template('users.html', users=users)

@app.route('/admin/users/add', methods=['POST'])
@admin_required
def add_user():
    username = request.form.get('username','').strip()
    password = request.form.get('password','')
    role     = request.form.get('role','viewer')
    if not username or not password:
        flash('Username and password are required.', 'danger')
        return redirect(url_for('manage_users'))
    conn = get_db()
    cur  = conn.cursor()
    try:
        cur.execute("INSERT INTO users (username, password_hash, role) VALUES (?,?,?)",
                    (username, generate_password_hash(password), role))
        conn.commit()
        flash(f"User '{username}' created successfully.", 'success')
    except sqlite3.IntegrityError:
        flash('Username already exists.', 'danger')
    conn.close()
    return redirect(url_for('manage_users'))

@app.route('/admin/users/delete/<int:uid>', methods=['POST'])
@admin_required
def delete_user(uid):
    if uid == session['user_id']:
        flash("You can't delete your own account.", 'danger')
        return redirect(url_for('manage_users'))
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("DELETE FROM users WHERE id=?", (uid,))
    conn.commit()
    conn.close()
    flash('User deleted.', 'success')
    return redirect(url_for('manage_users'))

@app.route('/admin/change-password', methods=['POST'])
@login_required
def change_password():
    current  = request.form.get('current_password','')
    new_pass = request.form.get('new_password','')
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("SELECT * FROM users WHERE id=?", (session['user_id'],))
    user = cur.fetchone()
    if not check_password_hash(user['password_hash'], current):
        flash('Current password is incorrect.', 'danger')
    elif len(new_pass) < 6:
        flash('New password must be at least 6 characters.', 'danger')
    else:
        cur.execute("UPDATE users SET password_hash=? WHERE id=?",
                    (generate_password_hash(new_pass), session['user_id']))
        conn.commit()
        flash('Password changed successfully.', 'success')
    conn.close()
    return redirect(url_for('manage_users') if session['role'] == 'admin' else url_for('report'))

# ── MAIN ROUTES ───────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def index():
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("SELECT name FROM subjects ORDER BY name")
    subjects = [r['name'] for r in cur.fetchall()]
    cur.execute("SELECT COUNT(*) FROM students")
    student_count = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM attendance_records")
    record_count = cur.fetchone()[0]
    conn.close()
    return render_template('index.html', subjects=subjects,
                           student_count=student_count, record_count=record_count)

@app.route('/upload', methods=['POST'])
@admin_required
def upload():
    subject      = request.form.get('subject', '').strip()
    new_subject  = request.form.get('new_subject', '').strip()
    subject_name = new_subject if new_subject else subject

    if not subject_name:
        flash('Please select or enter a subject name.', 'danger')
        return redirect(url_for('index'))
    if 'file' not in request.files or request.files['file'].filename == '':
        flash('No file selected.', 'danger')
        return redirect(url_for('index'))

    f = request.files['file']
    if not allowed_file(f.filename):
        flash('Only .xlsx / .xls files are allowed.', 'danger')
        return redirect(url_for('index'))

    filename = secure_filename(f.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    f.save(filepath)

    try:
        res = parse_and_store(filepath, subject_name)
        msg = f"✅ Done! {res['inserted']} new records, {res['updated']} updated."
        if res['errors']:
            msg += f" ⚠️ {len(res['errors'])} warnings."
            for e in res['errors'][:5]:
                print('WARN:', e)
        flash(msg, 'success')
    except Exception as e:
        flash(f'Error processing file: {e}', 'danger')

    return redirect(url_for('index'))

@app.route('/report')
@login_required
def report():
    subject_filter = request.args.get('subject', '')
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("SELECT name FROM subjects ORDER BY name")
    subjects = [r['name'] for r in cur.fetchall()]

    if subject_filter:
        cur.execute("""
            SELECT s.regno, s.name,
                   SUM(CASE WHEN ar.status='P' THEN 1 ELSE 0 END) AS present,
                   COUNT(ar.id) AS total,
                   ROUND(SUM(CASE WHEN ar.status='P' THEN 1 ELSE 0 END)*100.0/COUNT(ar.id),2) AS percentage
            FROM students s
            JOIN attendance_records ar ON ar.student_id = s.id
            JOIN subjects sub ON sub.id = ar.subject_id
            WHERE sub.name = ?
            GROUP BY s.id, s.regno, s.name
            ORDER BY s.regno
        """, (subject_filter,))
    else:
        cur.execute("""
            SELECT s.regno, s.name,
                   SUM(CASE WHEN ar.status='P' THEN 1 ELSE 0 END) AS present,
                   COUNT(ar.id) AS total,
                   ROUND(SUM(CASE WHEN ar.status='P' THEN 1 ELSE 0 END)*100.0/COUNT(ar.id),2) AS percentage
            FROM students s
            JOIN attendance_records ar ON ar.student_id = s.id
            GROUP BY s.id, s.regno, s.name
            ORDER BY s.regno
        """)

    rows = cur.fetchall()
    conn.close()
    return render_template('report.html', rows=rows, subjects=subjects,
                           subject_filter=subject_filter)

@app.route('/api/report')
@login_required
def api_report():
    subject = request.args.get('subject', '')
    conn = get_db()
    cur  = conn.cursor()
    if subject:
        cur.execute("""
            SELECT s.regno, s.name, sub.name AS subject,
                   SUM(CASE WHEN ar.status='P' THEN 1 ELSE 0 END) AS present,
                   COUNT(ar.id) AS total,
                   ROUND(SUM(CASE WHEN ar.status='P' THEN 1 ELSE 0 END)*100.0/COUNT(ar.id),2) AS percentage
            FROM students s
            JOIN attendance_records ar ON ar.student_id = s.id
            JOIN subjects sub ON sub.id = ar.subject_id
            WHERE sub.name=?
            GROUP BY s.id, s.regno, s.name, sub.name
            ORDER BY s.regno
        """, (subject,))
    else:
        cur.execute("""
            SELECT s.regno, s.name, sub.name AS subject,
                   SUM(CASE WHEN ar.status='P' THEN 1 ELSE 0 END) AS present,
                   COUNT(ar.id) AS total,
                   ROUND(SUM(CASE WHEN ar.status='P' THEN 1 ELSE 0 END)*100.0/COUNT(ar.id),2) AS percentage
            FROM students s
            JOIN attendance_records ar ON ar.student_id = s.id
            JOIN subjects sub ON sub.id = ar.subject_id
            GROUP BY s.id, s.regno, s.name, sub.name
            ORDER BY s.regno
        """)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify(rows)

if __name__ == '__main__':
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)
